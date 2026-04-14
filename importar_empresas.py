#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para importar empresas do arquivo XLSX para o banco bancodivan
"""

import re
import mysql.connector
from openpyxl import load_workbook

# ============================================
# CONFIGURAÇÕES DO BANCO
# ============================================
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': 'gil312452',
    'database': 'bancodivan',
    'charset': 'utf8mb4'
}

ARQUIVO_XLSX = 'C:/Users/gcsan/Downloads/empresas.xlsx'  # ← ALTERE AQUI

# ============================================
# FUNÇÕES AUXILIARES
# ============================================

def limpar_cnpj(cnpj):
    """Remove formatação do CNPJ"""
    if not cnpj:
        return None
    cnpj_str = str(cnpj).replace('.', '').replace('/', '').replace('-', '').strip()
    # Remove decimais se vier como float (ex: 30460209000107.0)
    cnpj_str = cnpj_str.split('.')[0]
    if len(cnpj_str) == 14:
        return f"{cnpj_str[:2]}.{cnpj_str[2:5]}.{cnpj_str[5:8]}/{cnpj_str[8:12]}-{cnpj_str[12:]}"
    return cnpj_str if cnpj_str else None

def limpar_telefone(fone):
    """Limpa e formata telefone"""
    if not fone:
        return None
    fone_str = str(fone)
    # Remover tudo que não é número
    numeros = re.sub(r'\D', '', fone_str)
    if len(numeros) == 11:
        return f"({numeros[:2]}) {numeros[2:7]}-{numeros[7:]}"
    elif len(numeros) == 10:
        return f"({numeros[:2]}) {numeros[2:6]}-{numeros[6:]}"
    return fone_str[:20] if fone_str else None

def limpar_texto(texto, limite=100):
    """Limpa e trunca texto"""
    if not texto:
        return None
    return str(texto).strip()[:limite]

# ============================================
# IMPORTAÇÃO
# ============================================

def importar():
    print("=" * 50)
    print("🏢 IMPORTAÇÃO DE EMPRESAS")
    print("=" * 50)

    # Carregar Excel
    wb = load_workbook(ARQUIVO_XLSX, read_only=True)
    ws = wb.active

    # Conectar ao banco
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    inseridos = 0
    ignorados = 0
    erros = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            ativo    = row[0]   # Ativo?
            nome     = row[1]   # Nome
            cnpj_raw = row[3]   # CNPJ
            fone_raw = row[8]   # Fone

            # Pular se não tiver nome
            if not nome or str(nome).strip() == '':
                print(f"  ⚠️ Linha {i}: sem nome — ignorado")
                ignorados += 1
                continue

            nome_empresa = limpar_texto(nome, 100)
            cnpj = limpar_cnpj(cnpj_raw)
            celular = limpar_telefone(fone_raw)
            contato = nome_empresa  # usar nome como contato padrão

            # Verificar se já existe pelo nome
            cursor.execute(
                "SELECT id FROM empresas WHERE nome_empresa = %s",
                (nome_empresa,)
            )
            if cursor.fetchone():
                print(f"  ⏭️ Linha {i}: '{nome_empresa}' já existe — ignorado")
                ignorados += 1
                continue

            # Verificar CNPJ duplicado
            if cnpj:
                cursor.execute(
                    "SELECT id FROM empresas WHERE cnpj = %s",
                    (cnpj,)
                )
                if cursor.fetchone():
                    print(f"  ⏭️ Linha {i}: CNPJ {cnpj} já existe — ignorado")
                    ignorados += 1
                    continue

            # Inserir
            cursor.execute("""
                INSERT INTO empresas (nome_empresa, cnpj, celular, contato, autoriza_todos_jantar)
                VALUES (%s, %s, %s, %s, 0)
            """, (nome_empresa, cnpj, celular, contato))

            conn.commit()
            inseridos += 1
            print(f"  ✅ Linha {i}: '{nome_empresa}' inserida")

        except Exception as e:
            conn.rollback()
            erros += 1
            print(f"  ❌ Linha {i}: ERRO — {e}")

    cursor.close()
    conn.close()

    print()
    print("=" * 50)
    print(f"✅ Inseridas : {inseridos}")
    print(f"⏭️ Ignoradas : {ignorados}")
    print(f"❌ Erros     : {erros}")
    print("=" * 50)

if __name__ == '__main__':
    importar()